"""
XLSX Tax Report Generator — with USD valuations.

Produces a workbook with:
  - Summary tab (totals + USD volumes per exchange)
  - One tab per exchange: Trades, Orders, Deposits, Withdrawals, Pool Activity
  - All monetary values include USD equivalents
"""
import logging
import os
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

logger = logging.getLogger("tax-collector.export")

EXPORT_DIR = os.environ.get("TAX_EXPORT_DIR", "/data/exports")

HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
SECTION_FONT = Font(name="Arial", bold=True, size=12, color="2F5496")
DATA_FONT = Font(name="Arial", size=10)
USD_FONT = Font(name="Arial", size=10, color="006100")
MISSING_FONT = Font(name="Arial", size=10, color="CC0000", italic=True)
CRYPTO_FMT = '#,##0.00000000'
USD_FMT = '$#,##0.00'
DATE_FMT = "YYYY-MM-DD HH:MM:SS"
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def _style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


async def _fetch_rows(session: AsyncSession, table: str, exchange: str, year: int | None,
                      date_col: str = "executed_at") -> list[dict]:
    where = "WHERE exchange = :ex"
    params: dict = {"ex": exchange}
    if year:
        where += f" AND EXTRACT(YEAR FROM {date_col}) = :year"
        params["year"] = year
    result = await session.execute(
        text(f"SELECT * FROM tax.{table} {where} ORDER BY {date_col} ASC"), params)
    cols = list(result.keys())
    return [dict(zip(cols, row)) for row in result.fetchall()]


def _write_cell(ws, row, col, value, is_usd=False, is_crypto=False, is_date=False):
    """Write a cell with appropriate formatting."""
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER

    if is_date and isinstance(value, datetime):
        cell.value = value
        cell.number_format = DATE_FMT
        cell.font = DATA_FONT
    elif is_usd:
        try:
            fval = float(value) if value else 0
            cell.value = fval
            cell.number_format = USD_FMT
            cell.font = USD_FONT if fval > 0 else MISSING_FONT
        except (ValueError, TypeError):
            cell.value = "N/A"
            cell.font = MISSING_FONT
    elif is_crypto:
        try:
            cell.value = float(value) if value else 0
            cell.number_format = CRYPTO_FMT
            cell.font = DATA_FONT
        except (ValueError, TypeError):
            cell.value = str(value) if value else ""
            cell.font = DATA_FONT
    else:
        cell.value = str(value) if value is not None else ""
        cell.font = DATA_FONT


def _write_section(ws, start_row: int, title: str, headers: list[str],
                   rows: list[dict], col_map: list[str],
                   usd_cols: set[str] | None = None,
                   crypto_cols: set[str] | None = None,
                   date_cols: set[str] | None = None) -> int:
    usd_cols = usd_cols or set()
    crypto_cols = crypto_cols or set()
    date_cols = date_cols or set()
    row = start_row

    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    row += 1

    if not rows:
        ws.cell(row=row, column=1, value="No data").font = Font(
            name="Arial", size=10, italic=True, color="999999")
        return row + 2

    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    for r in rows:
        for ci, key in enumerate(col_map, 1):
            val = r.get(key, "")
            _write_cell(ws, row, ci, val,
                        is_usd=(key in usd_cols),
                        is_crypto=(key in crypto_cols),
                        is_date=(key in date_cols))
        row += 1

    return row + 1


async def generate_tax_xlsx(session: AsyncSession, year: int | None = None) -> str:
    os.makedirs(EXPORT_DIR, exist_ok=True)
    year_label = str(year) if year else "all_years"
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"tax_report_{year_label}_{ts}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb = Workbook()

    # Get exchanges with data
    result = await session.execute(text(
        "SELECT DISTINCT exchange FROM tax.trades "
        "UNION SELECT DISTINCT exchange FROM tax.deposits "
        "UNION SELECT DISTINCT exchange FROM tax.withdrawals "
        "ORDER BY 1"))
    exchanges = [row[0] for row in result.fetchall()]

    if not exchanges:
        ws = wb.active
        ws.title = "No Data"
        ws.cell(row=1, column=1, value="No trading data found. Run a sync first.")
        wb.save(filepath)
        return filepath

    # ── Summary Tab ───────────────────────────────────────────────────────

    ws = wb.active
    ws.title = "Summary"

    ws.cell(row=1, column=1, value=f"Tax Report — {year_label}").font = Font(
        name="Arial", bold=True, size=14, color="2F5496")
    ws.cell(row=2, column=1,
            value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = Font(
        name="Arial", size=10, color="666666")
    ws.cell(row=3, column=1,
            value="USD values sourced from CoinGecko historical daily prices").font = Font(
        name="Arial", size=9, color="999999")

    sum_headers = ["Exchange", "Total Trades", "Buys", "Sells",
                   "Trade Volume (USD)", "Trading Fees (USD)",
                   "Deposits", "Deposits (USD)",
                   "Withdrawals", "Withdrawals (USD)", "Withdrawal Fees (USD)",
                   "Pool Events"]
    sr = 5
    for ci, h in enumerate(sum_headers, 1):
        ws.cell(row=sr, column=ci, value=h)
    _style_header_row(ws, sr, len(sum_headers))
    sr += 1

    for ex in exchanges:
        p: dict = {"ex": ex}
        yf = ""
        if year:
            yf = "AND EXTRACT(YEAR FROM executed_at) = :year"
            p["year"] = year
        yf_dep = yf.replace("executed_at", "confirmed_at")

        r = await session.execute(text(
            f"SELECT COUNT(*) FROM tax.trades WHERE exchange=:ex {yf}"), p)
        total = r.scalar() or 0
        r = await session.execute(text(
            f"SELECT COUNT(*) FROM tax.trades WHERE exchange=:ex AND side='buy' {yf}"), p)
        buys = r.scalar() or 0
        r = await session.execute(text(
            f"SELECT COUNT(*) FROM tax.trades WHERE exchange=:ex AND side='sell' {yf}"), p)
        sells = r.scalar() or 0
        r = await session.execute(text(
            f"SELECT COALESCE(SUM(total_usd),0) FROM tax.trades WHERE exchange=:ex {yf}"), p)
        vol_usd = float(r.scalar() or 0)
        r = await session.execute(text(
            f"SELECT COALESCE(SUM(fee_usd),0) FROM tax.trades WHERE exchange=:ex {yf}"), p)
        fees_usd = float(r.scalar() or 0)
        r = await session.execute(text(
            f"SELECT COUNT(*), COALESCE(SUM(amount_usd),0) FROM tax.deposits WHERE exchange=:ex {yf_dep}"), p)
        dep_row = r.fetchone()
        dep_count = dep_row[0] if dep_row else 0
        dep_usd = float(dep_row[1]) if dep_row else 0
        r = await session.execute(text(
            f"SELECT COUNT(*), COALESCE(SUM(amount_usd),0), COALESCE(SUM(fee_usd),0) "
            f"FROM tax.withdrawals WHERE exchange=:ex {yf_dep}"), p)
        wd_row = r.fetchone()
        wd_count = wd_row[0] if wd_row else 0
        wd_usd = float(wd_row[1]) if wd_row else 0
        wd_fee_usd = float(wd_row[2]) if wd_row else 0
        r = await session.execute(text(
            f"SELECT COUNT(*) FROM tax.pool_activity WHERE exchange=:ex {yf}"), p)
        pool_count = r.scalar() or 0

        vals = [ex.upper(), total, buys, sells, vol_usd, fees_usd,
                dep_count, dep_usd, wd_count, wd_usd, wd_fee_usd, pool_count]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=sr, column=ci, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if ci in (5, 6, 8, 10, 11):
                cell.number_format = USD_FMT
                cell.font = USD_FONT
        sr += 1

    # Missing USD warning
    r = await session.execute(text("SELECT COUNT(*) FROM tax.trades WHERE total_usd IS NULL"))
    t_miss = r.scalar() or 0
    r = await session.execute(text("SELECT COUNT(*) FROM tax.deposits WHERE amount_usd IS NULL"))
    d_miss = r.scalar() or 0
    r = await session.execute(text("SELECT COUNT(*) FROM tax.withdrawals WHERE amount_usd IS NULL"))
    w_miss = r.scalar() or 0
    if t_miss + d_miss + w_miss > 0:
        sr += 1
        ws.cell(row=sr, column=1,
                value=f"Warning: {t_miss} trades, {d_miss} deposits, {w_miss} withdrawals "
                      f"missing USD values. Run /backfill-usd to resolve.").font = Font(
            name="Arial", size=10, color="CC0000", bold=True)

    _auto_width(ws)

    # ── Per-Exchange Tabs ─────────────────────────────────────────────────

    trade_usd_cols = {"price_usd", "quantity_usd", "total_usd", "fee_usd", "base_price_usd", "quote_price_usd"}
    trade_crypto_cols = {"price", "quantity", "total", "fee"}
    trade_date_cols = {"executed_at"}

    order_usd_cols = {"price_usd", "total_usd"}
    order_crypto_cols = {"price", "quantity", "executed_qty"}
    order_date_cols = {"created_at_ex", "updated_at_ex"}

    dep_usd_cols = {"asset_price_usd", "amount_usd"}
    dep_crypto_cols = {"amount"}
    dep_date_cols = {"confirmed_at"}

    wd_usd_cols = {"asset_price_usd", "amount_usd", "fee_usd"}
    wd_crypto_cols = {"amount", "fee"}
    wd_date_cols = {"confirmed_at"}

    pool_usd_cols = {"amount_in_usd", "amount_out_usd", "fee_usd"}
    pool_crypto_cols = {"amount_in", "amount_out", "fee"}
    pool_date_cols = {"executed_at"}

    for ex in exchanges:
        ws_ex = wb.create_sheet(title=ex.upper()[:31])
        row = 1

        ws_ex.cell(row=row, column=1, value=f"{ex.upper()} — Trading Activity").font = Font(
            name="Arial", bold=True, size=14, color="2F5496")
        row += 2

        # Trades
        trades = await _fetch_rows(session, "trades", ex, year, "executed_at")
        trade_headers = ["Date", "Market", "Base", "Quote", "Side",
                         "Price", "Quantity", "Total", "Fee", "Fee Asset",
                         "Price (USD)", "Qty (USD)", "Total (USD)", "Fee (USD)",
                         "Base Rate", "Quote Rate", "Exchange ID"]
        trade_cols = ["executed_at", "market", "base_asset", "quote_asset", "side",
                      "price", "quantity", "total", "fee", "fee_asset",
                      "price_usd", "quantity_usd", "total_usd", "fee_usd",
                      "base_price_usd", "quote_price_usd", "exchange_id"]
        row = _write_section(ws_ex, row, "Trades", trade_headers, trades, trade_cols,
                             usd_cols=trade_usd_cols, crypto_cols=trade_crypto_cols,
                             date_cols=trade_date_cols)

        # Orders
        orders = await _fetch_rows(session, "orders", ex, year, "created_at_ex")
        order_headers = ["Date", "Market", "Base", "Quote", "Side", "Type",
                         "Price", "Quantity", "Executed", "Status",
                         "Price (USD)", "Total (USD)",
                         "Updated", "Exchange ID"]
        order_cols = ["created_at_ex", "market", "base_asset", "quote_asset", "side", "order_type",
                      "price", "quantity", "executed_qty", "status",
                      "price_usd", "total_usd",
                      "updated_at_ex", "exchange_id"]
        row = _write_section(ws_ex, row, "Orders", order_headers, orders, order_cols,
                             usd_cols=order_usd_cols, crypto_cols=order_crypto_cols,
                             date_cols=order_date_cols)

        # Deposits
        deposits = await _fetch_rows(session, "deposits", ex, year, "confirmed_at")
        dep_headers = ["Date", "Asset", "Amount", "Asset Price (USD)", "Amount (USD)",
                       "Network", "TX Hash", "Address", "Status"]
        dep_cols = ["confirmed_at", "asset", "amount", "asset_price_usd", "amount_usd",
                    "network", "tx_hash", "address", "status"]
        row = _write_section(ws_ex, row, "Deposits", dep_headers, deposits, dep_cols,
                             usd_cols=dep_usd_cols, crypto_cols=dep_crypto_cols,
                             date_cols=dep_date_cols)

        # Withdrawals
        withdrawals = await _fetch_rows(session, "withdrawals", ex, year, "confirmed_at")
        wd_headers = ["Date", "Asset", "Amount", "Fee",
                      "Asset Price (USD)", "Amount (USD)", "Fee (USD)",
                      "Network", "TX Hash", "Address", "Status"]
        wd_cols = ["confirmed_at", "asset", "amount", "fee",
                   "asset_price_usd", "amount_usd", "fee_usd",
                   "network", "tx_hash", "address", "status"]
        row = _write_section(ws_ex, row, "Withdrawals", wd_headers, withdrawals, wd_cols,
                             usd_cols=wd_usd_cols, crypto_cols=wd_crypto_cols,
                             date_cols=wd_date_cols)

        # Pool Activity
        pools = await _fetch_rows(session, "pool_activity", ex, year, "executed_at")
        pool_headers = ["Date", "Pool", "Action",
                        "Asset In", "Amount In", "In (USD)",
                        "Asset Out", "Amount Out", "Out (USD)",
                        "Fee", "Fee Asset", "Fee (USD)"]
        pool_cols = ["executed_at", "pool_name", "action",
                     "asset_in", "amount_in", "amount_in_usd",
                     "asset_out", "amount_out", "amount_out_usd",
                     "fee", "fee_asset", "fee_usd"]
        row = _write_section(ws_ex, row, "Pool Activity", pool_headers, pools, pool_cols,
                             usd_cols=pool_usd_cols, crypto_cols=pool_crypto_cols,
                             date_cols=pool_date_cols)

        _auto_width(ws_ex)

    wb.save(filepath)
    logger.info(f"Tax report exported: {filepath}")
    return filepath