"""
Full Tax Report XLSX Generator — accountant-ready output.

Tabs:
  1. Summary         — high-level numbers, Schedule D totals
  2. Form 8949 (ST)  — short-term capital gains/losses (Box B)
  3. Form 8949 (LT)  — long-term capital gains/losses (Box D)
  4. Income Schedule  — staking rewards, airdrops (ordinary income)
  5. Transfer Recon   — matched cross-exchange transfers (non-taxable)
  6. Fee Summary      — deductible trading fees by exchange
  7. Lot Inventory    — all acquisition lots with remaining quantities
  8+  Per-exchange raw data tabs (from original export)

This function is called via GET /export/tax-report?year=2025
"""
import logging
import os
from datetime import datetime, timezone
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

logger = logging.getLogger("tax-collector.export")

EXPORT_DIR = os.environ.get("TAX_EXPORT_DIR", "/data/exports")

# Styles
H_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
H_FILL = PatternFill("solid", fgColor="2F5496")
H2_FONT = Font(name="Arial", bold=True, size=12, color="2F5496")
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="2F5496")
DATA_FONT = Font(name="Arial", size=10)
USD_FONT = Font(name="Arial", size=10, color="006100")
LOSS_FONT = Font(name="Arial", size=10, color="CC0000")
DIM_FONT = Font(name="Arial", size=10, color="999999", italic=True)
NOTE_FONT = Font(name="Arial", size=9, color="666666")
BOLD_FONT = Font(name="Arial", bold=True, size=10)
USD_FMT = '$#,##0.00'
CRYPTO_FMT = '#,##0.00000000'
DATE_FMT = "YYYY-MM-DD HH:MM:SS"
BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))

D = Decimal


def _hdr(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = H_FONT
        c.fill = H_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")


def _auto(ws, mn=10, mx=35):
    for col_cells in ws.columns:
        ml = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(ml + 2, mn), mx)


def _usd_cell(ws, row, col, val):
    c = ws.cell(row=row, column=col)
    try:
        fv = float(val) if val else 0
        c.value = fv
        c.number_format = USD_FMT
        c.font = LOSS_FONT if fv < 0 else USD_FONT
    except (ValueError, TypeError):
        c.value = "N/A"
        c.font = DIM_FONT
    c.border = BORDER


def _data_cell(ws, row, col, val, is_date=False):
    c = ws.cell(row=row, column=col)
    c.border = BORDER
    c.font = DATA_FONT
    if is_date and isinstance(val, datetime):
        c.value = val
        c.number_format = DATE_FMT
    else:
        c.value = str(val) if val is not None else ""


async def generate_full_tax_report(session: AsyncSession, year: int) -> str:
    """Generate the full accountant-ready XLSX tax report."""
    os.makedirs(EXPORT_DIR, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"tax_report_{year}_full_{ts}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb = Workbook()

    await _build_summary(wb, session, year)
    await _build_form_8949(wb, session, year, "short", "Form 8949 (ST)")
    await _build_form_8949(wb, session, year, "long", "Form 8949 (LT)")
    await _build_income_schedule(wb, session, year)
    await _build_transfer_recon(wb, session)
    await _build_fee_summary(wb, session, year)
    await _build_lot_inventory(wb, session)
    await _build_raw_trades(wb, session, year)

    wb.save(filepath)
    logger.info(f"Full tax report exported: {filepath}")
    return filepath


# ── Tab 1: Summary ────────────────────────────────────────────────────────

async def _build_summary(wb, session, year):
    ws = wb.active
    ws.title = "Summary"

    ws.cell(row=1, column=1, value=f"Cryptocurrency Tax Report — {year}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = NOTE_FONT
    ws.cell(row=3, column=1, value="Cost basis method: FIFO (First In, First Out)").font = NOTE_FONT

    # Schedule D summary
    ws.cell(row=5, column=1, value="SCHEDULE D SUMMARY").font = H2_FONT
    r = 6
    _hdr(ws, r, ["Category", "Proceeds (USD)", "Cost Basis (USD)", "Net Gain/Loss (USD)"])
    r += 1

    res = await session.execute(text("""
        SELECT
            COALESCE(SUM(CASE WHEN term='short' THEN proceeds END), 0),
            COALESCE(SUM(CASE WHEN term='short' THEN cost_basis END), 0),
            COALESCE(SUM(CASE WHEN term='short' THEN gain_loss END), 0),
            COALESCE(SUM(CASE WHEN term='long' THEN proceeds END), 0),
            COALESCE(SUM(CASE WHEN term='long' THEN cost_basis END), 0),
            COALESCE(SUM(CASE WHEN term='long' THEN gain_loss END), 0),
            COUNT(*)
        FROM tax.form_8949 WHERE tax_year = :y
    """), {"y": year})
    d = res.fetchone()

    for label, proc, cost, gl in [
        ("Short-Term (< 1 year)", d[0], d[1], d[2]),
        ("Long-Term (≥ 1 year)", d[3], d[4], d[5]),
        ("TOTAL", float(d[0]) + float(d[3]), float(d[1]) + float(d[4]), float(d[2]) + float(d[5])),
    ]:
        _data_cell(ws, r, 1, label)
        ws.cell(row=r, column=1).font = BOLD_FONT if label == "TOTAL" else DATA_FONT
        _usd_cell(ws, r, 2, proc)
        _usd_cell(ws, r, 3, cost)
        _usd_cell(ws, r, 4, gl)
        r += 1

    # Ordinary income summary
    r += 1
    ws.cell(row=r, column=1, value="ORDINARY INCOME (Staking, Rewards)").font = H2_FONT
    r += 1
    _hdr(ws, r, ["Income Type", "Count", "Total (USD)"])
    r += 1

    res = await session.execute(text("""
        SELECT income_type, COUNT(*), COALESCE(SUM(amount_usd), 0)
        FROM tax.income_events
        WHERE EXTRACT(YEAR FROM received_at) = :y
        GROUP BY income_type ORDER BY income_type
    """), {"y": year})
    income_total = D("0")
    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0].replace("_", " ").title())
        _data_cell(ws, r, 2, row_data[1])
        _usd_cell(ws, r, 3, row_data[2])
        income_total += D(str(row_data[2]))
        r += 1
    _data_cell(ws, r, 1, "TOTAL INCOME")
    ws.cell(row=r, column=1).font = BOLD_FONT
    _usd_cell(ws, r, 3, str(income_total))
    r += 2

    # Fee summary
    ws.cell(row=r, column=1, value="DEDUCTIBLE TRADING FEES").font = H2_FONT
    r += 1
    res = await session.execute(text("""
        SELECT exchange, COALESCE(SUM(fee_usd), 0)
        FROM tax.trades
        WHERE fee_usd > 0 AND EXTRACT(YEAR FROM executed_at) = :y
        GROUP BY exchange
    """), {"y": year})
    fee_total = D("0")
    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0].upper())
        _usd_cell(ws, r, 2, row_data[1])
        fee_total += D(str(row_data[1]))
        r += 1
    _data_cell(ws, r, 1, "TOTAL FEES")
    ws.cell(row=r, column=1).font = BOLD_FONT
    _usd_cell(ws, r, 2, str(fee_total))
    r += 2

    # Transfer summary
    res = await session.execute(text("SELECT COUNT(*) FROM tax.transfer_matches"))
    tm_count = res.scalar() or 0
    ws.cell(row=r, column=1, value=f"Cross-exchange transfers matched: {tm_count} (non-taxable)").font = NOTE_FONT

    _auto(ws)


# ── Tab 2/3: Form 8949 ───────────────────────────────────────────────────

async def _build_form_8949(wb, session, year, term, tab_name):
    ws = wb.create_sheet(title=tab_name)

    box = "B" if term == "short" else "D"
    label = "Short-Term" if term == "short" else "Long-Term"

    ws.cell(row=1, column=1, value=f"Form 8949 — {label} Capital Gains and Losses").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Tax Year: {year} | Box {box} (basis NOT reported to IRS)").font = NOTE_FONT

    r = 4
    headers = [
        "(a) Description", "(b) Date Acquired", "(c) Date Sold",
        "(d) Proceeds", "(e) Cost Basis",
        "(f) Code", "(g) Adjustment",
        "(h) Gain or Loss", "Asset", "Exchange", "Holding Days"
    ]
    _hdr(ws, r, headers)
    r += 1

    res = await session.execute(text("""
        SELECT description, date_acquired, date_sold, proceeds, cost_basis,
               adjustment_code, adjustment_amount, gain_loss, asset, exchange, holding_days
        FROM tax.form_8949
        WHERE tax_year = :y AND term = :t
        ORDER BY date_sold, asset
    """), {"y": year, "t": term})

    total_proceeds = D("0")
    total_cost = D("0")
    total_gl = D("0")
    count = 0

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0])
        _data_cell(ws, r, 2, row_data[1])
        _data_cell(ws, r, 3, row_data[2])
        _usd_cell(ws, r, 4, row_data[3])
        _usd_cell(ws, r, 5, row_data[4])
        _data_cell(ws, r, 6, row_data[5] or "")
        _usd_cell(ws, r, 7, row_data[6])
        _usd_cell(ws, r, 8, row_data[7])
        _data_cell(ws, r, 9, row_data[8])
        _data_cell(ws, r, 10, row_data[9])
        _data_cell(ws, r, 11, row_data[10])
        total_proceeds += D(str(row_data[3] or 0))
        total_cost += D(str(row_data[4] or 0))
        total_gl += D(str(row_data[7] or 0))
        count += 1
        r += 1

    # Totals row
    r += 1
    ws.cell(row=r, column=1, value=f"TOTALS ({count} disposals)").font = BOLD_FONT
    _usd_cell(ws, r, 4, str(total_proceeds))
    ws.cell(row=r, column=4).font = Font(name="Arial", bold=True, size=10, color="006100")
    _usd_cell(ws, r, 5, str(total_cost))
    ws.cell(row=r, column=5).font = Font(name="Arial", bold=True, size=10, color="006100")
    _usd_cell(ws, r, 8, str(total_gl))
    ws.cell(row=r, column=8).font = Font(name="Arial", bold=True, size=10,
                                          color="CC0000" if total_gl < 0 else "006100")

    _auto(ws)


# ── Tab 4: Income Schedule ────────────────────────────────────────────────

async def _build_income_schedule(wb, session, year):
    ws = wb.create_sheet(title="Income Schedule")

    ws.cell(row=1, column=1, value=f"Ordinary Income Schedule — {year}").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Staking rewards and airdrops are taxed as ordinary income at FMV when received").font = NOTE_FONT

    r = 4
    _hdr(ws, r, ["Date Received", "Type", "Asset", "Amount", "FMV (USD)", "Exchange", "Description"])
    r += 1

    res = await session.execute(text("""
        SELECT received_at, income_type, asset, amount, amount_usd, exchange, description
        FROM tax.income_events
        WHERE EXTRACT(YEAR FROM received_at) = :y
        ORDER BY received_at ASC
    """), {"y": year})

    total = D("0")
    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0], is_date=True)
        _data_cell(ws, r, 2, row_data[1].replace("_", " ").title())
        _data_cell(ws, r, 3, row_data[2])
        _data_cell(ws, r, 4, row_data[3])
        ws.cell(row=r, column=4).number_format = CRYPTO_FMT
        _usd_cell(ws, r, 5, row_data[4])
        _data_cell(ws, r, 6, row_data[5])
        _data_cell(ws, r, 7, row_data[6])
        total += D(str(row_data[4] or 0))
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="TOTAL ORDINARY INCOME").font = BOLD_FONT
    _usd_cell(ws, r, 5, str(total))

    _auto(ws)


# ── Tab 5: Transfer Reconciliation ───────────────────────────────────────

async def _build_transfer_recon(wb, session):
    ws = wb.create_sheet(title="Transfer Recon")

    ws.cell(row=1, column=1, value="Cross-Exchange Transfer Reconciliation").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Matched transfers are NOT taxable events — cost basis carries over").font = NOTE_FONT

    r = 4
    _hdr(ws, r, ["Date", "Asset", "Amount", "From Exchange", "To Exchange",
                  "TX Hash", "Confidence", "Cost Basis (USD)"])
    r += 1

    res = await session.execute(text("""
        SELECT transferred_at, asset, amount, from_exchange, to_exchange,
               tx_hash, match_confidence, cost_basis_usd
        FROM tax.transfer_matches
        ORDER BY transferred_at ASC
    """))

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0], is_date=True)
        _data_cell(ws, r, 2, row_data[1])
        _data_cell(ws, r, 3, row_data[2])
        ws.cell(row=r, column=3).number_format = CRYPTO_FMT
        _data_cell(ws, r, 4, (row_data[3] or "").upper())
        _data_cell(ws, r, 5, (row_data[4] or "").upper())
        _data_cell(ws, r, 6, row_data[5] or "")
        _data_cell(ws, r, 7, row_data[6] or "")
        _usd_cell(ws, r, 8, row_data[7])
        r += 1

    if r == 5:
        _data_cell(ws, r, 1, "No matched transfers found")
        ws.cell(row=r, column=1).font = DIM_FONT

    _auto(ws)


# ── Tab 6: Fee Summary ───────────────────────────────────────────────────

async def _build_fee_summary(wb, session, year):
    ws = wb.create_sheet(title="Fee Summary")

    ws.cell(row=1, column=1, value=f"Trading Fee Summary — {year}").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Trading fees may be deductible — consult your tax professional").font = NOTE_FONT

    r = 4
    _hdr(ws, r, ["Exchange", "Total Trades", "Trades with Fees", "Total Fees (USD)"])
    r += 1

    res = await session.execute(text("""
        SELECT exchange,
               COUNT(*),
               COUNT(CASE WHEN fee_usd > 0 THEN 1 END),
               COALESCE(SUM(fee_usd), 0)
        FROM tax.trades
        WHERE EXTRACT(YEAR FROM executed_at) = :y
        GROUP BY exchange ORDER BY exchange
    """), {"y": year})

    grand_total = D("0")
    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0].upper())
        _data_cell(ws, r, 2, row_data[1])
        _data_cell(ws, r, 3, row_data[2])
        _usd_cell(ws, r, 4, row_data[3])
        grand_total += D(str(row_data[3]))
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="TOTAL DEDUCTIBLE FEES").font = BOLD_FONT
    _usd_cell(ws, r, 4, str(grand_total))

    # Per-asset fee breakdown
    r += 2
    ws.cell(row=r, column=1, value="Fees by Asset").font = H2_FONT
    r += 1
    _hdr(ws, r, ["Fee Asset", "Total Fees", "Total Fees (USD)"])
    r += 1

    res = await session.execute(text("""
        SELECT COALESCE(fee_asset, 'Unknown'), SUM(fee)::text, COALESCE(SUM(fee_usd), 0)
        FROM tax.trades
        WHERE fee > 0 AND EXTRACT(YEAR FROM executed_at) = :y
        GROUP BY fee_asset ORDER BY SUM(fee_usd) DESC NULLS LAST
    """), {"y": year})

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0])
        _data_cell(ws, r, 2, row_data[1])
        ws.cell(row=r, column=2).number_format = CRYPTO_FMT
        _usd_cell(ws, r, 3, row_data[2])
        r += 1

    _auto(ws)


# ── Tab 7: Lot Inventory ─────────────────────────────────────────────────

async def _build_lot_inventory(wb, session):
    ws = wb.create_sheet(title="Lot Inventory")

    ws.cell(row=1, column=1, value="FIFO Lot Inventory — Current Holdings").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Shows remaining acquisition lots and their cost basis").font = NOTE_FONT

    r = 4
    _hdr(ws, r, ["Asset", "Acquired", "Source", "Exchange",
                  "Original Qty", "Remaining", "Cost/Unit (USD)", "Total Cost (USD)"])
    r += 1

    res = await session.execute(text("""
        SELECT asset, acquired_at, source, exchange,
               quantity, remaining, cost_per_unit_usd, total_cost_usd
        FROM tax.lots
        WHERE remaining > 0
        ORDER BY asset, acquired_at
    """))

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0])
        _data_cell(ws, r, 2, row_data[1], is_date=True)
        _data_cell(ws, r, 3, (row_data[2] or "").replace("_", " ").title())
        _data_cell(ws, r, 4, (row_data[3] or "").upper())
        _data_cell(ws, r, 5, row_data[4])
        ws.cell(row=r, column=5).number_format = CRYPTO_FMT
        _data_cell(ws, r, 6, row_data[5])
        ws.cell(row=r, column=6).number_format = CRYPTO_FMT
        _usd_cell(ws, r, 7, row_data[6])
        _usd_cell(ws, r, 8, row_data[7])
        r += 1

    _auto(ws)


# ── Tab 8+: Raw trade data per exchange ──────────────────────────────────

async def _build_raw_trades(wb, session, year):
    """Add per-exchange raw trade data tabs."""
    res = await session.execute(text("""
        SELECT DISTINCT exchange FROM tax.trades
        WHERE EXTRACT(YEAR FROM executed_at) = :y
        ORDER BY 1
    """), {"y": year})
    exchanges = [row[0] for row in res.fetchall()]

    for ex in exchanges:
        tab_name = f"{ex.upper()} Trades"[:31]
        ws = wb.create_sheet(title=tab_name)

        ws.cell(row=1, column=1, value=f"{ex.upper()} — All Trades {year}").font = TITLE_FONT
        r = 3
        headers = ["Date", "Market", "Side", "Price", "Qty", "Total",
                   "Fee", "Fee Asset", "Total (USD)", "Fee (USD)", "Exchange ID"]
        _hdr(ws, r, headers)
        r += 1

        trades = await session.execute(text("""
            SELECT executed_at, market, side, price, quantity, total,
                   fee, fee_asset, total_usd, fee_usd, exchange_id
            FROM tax.trades
            WHERE exchange = :ex AND EXTRACT(YEAR FROM executed_at) = :y
            ORDER BY executed_at ASC
        """), {"ex": ex, "y": year})

        for t in trades.fetchall():
            _data_cell(ws, r, 1, t[0], is_date=True)
            _data_cell(ws, r, 2, t[1])
            c = ws.cell(row=r, column=3, value=(t[2] or "").upper())
            c.font = Font(name="Arial", size=10, bold=True,
                          color="006100" if t[2] == "buy" else "CC0000")
            _data_cell(ws, r, 4, t[3])
            ws.cell(row=r, column=4).number_format = CRYPTO_FMT
            _data_cell(ws, r, 5, t[4])
            ws.cell(row=r, column=5).number_format = CRYPTO_FMT
            _data_cell(ws, r, 6, t[5])
            ws.cell(row=r, column=6).number_format = CRYPTO_FMT
            _data_cell(ws, r, 7, t[6])
            ws.cell(row=r, column=7).number_format = CRYPTO_FMT
            _data_cell(ws, r, 8, t[7])
            _usd_cell(ws, r, 9, t[8])
            _usd_cell(ws, r, 10, t[9])
            _data_cell(ws, r, 11, t[10])
            r += 1

        _auto(ws)


# ══════════════════════════════════════════════════════════════════════════
# V4 TAX REPORT — reads from v4 tables
# ══════════════════════════════════════════════════════════════════════════

async def generate_full_tax_report_v4(session: AsyncSession, year: int,
                                       run_id: int = None) -> str:
    """Generate the full accountant-ready XLSX tax report from v4 tables."""
    os.makedirs(EXPORT_DIR, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"tax_report_v4_{year}_full_{ts}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb = Workbook()

    await _build_summary_v4(wb, session, year, run_id)
    await _build_form_8949_v4(wb, session, year, "short", "Form 8949 (ST)", run_id)
    await _build_form_8949_v4(wb, session, year, "long", "Form 8949 (LT)", run_id)
    await _build_income_schedule_v4(wb, session, year, run_id)
    await _build_transfer_recon_v4(wb, session, run_id)
    await _build_lot_inventory_v4(wb, session, run_id)
    await _build_exchange_pnl_tab(wb, session, year, run_id)
    await _build_funding_flows_tab(wb, session, year)
    await _build_exceptions_tab(wb, session, run_id)
    await _build_data_coverage_tab(wb, session)
    await _build_valuation_audit_tab(wb, session, year, run_id)
    await _build_run_manifest_tab(wb, session, run_id)
    await _build_filing_readiness(wb, session, year, run_id)

    wb.save(filepath)
    logger.info(f"V4 full tax report exported: {filepath}")
    return filepath


async def _build_summary_v4(wb, session, year, run_id):
    ws = wb.active
    ws.title = "Summary"

    ws.cell(row=1, column=1, value=f"Cryptocurrency Tax Report (v4) — {year}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = NOTE_FONT
    ws.cell(row=3, column=1, value="Cost basis method: FIFO (Wallet-Aware, Filing-Grade)").font = NOTE_FONT

    ws.cell(row=5, column=1, value="SCHEDULE D SUMMARY").font = H2_FONT
    r = 6
    _hdr(ws, r, ["Category", "Proceeds (USD)", "Cost Basis (USD)", "Net Gain/Loss (USD)"])
    r += 1

    run_filter = "AND run_id = :rid" if run_id else ""
    params = {"y": year}
    if run_id:
        params["rid"] = run_id

    res = await session.execute(text(f"""
        SELECT
            COALESCE(SUM(CASE WHEN term='short' THEN proceeds END), 0),
            COALESCE(SUM(CASE WHEN term='short' THEN cost_basis END), 0),
            COALESCE(SUM(CASE WHEN term='short' THEN gain_loss END), 0),
            COALESCE(SUM(CASE WHEN term='long' THEN proceeds END), 0),
            COALESCE(SUM(CASE WHEN term='long' THEN cost_basis END), 0),
            COALESCE(SUM(CASE WHEN term='long' THEN gain_loss END), 0),
            COUNT(*)
        FROM tax.form_8949_v4 WHERE tax_year = :y {run_filter}
    """), params)
    d = res.fetchone()

    for label, proc, cost, gl in [
        ("Short-Term (< 1 year)", d[0], d[1], d[2]),
        ("Long-Term (> 1 year)", d[3], d[4], d[5]),
        ("TOTAL", float(d[0]) + float(d[3]), float(d[1]) + float(d[4]), float(d[2]) + float(d[5])),
    ]:
        _data_cell(ws, r, 1, label)
        ws.cell(row=r, column=1).font = BOLD_FONT if label == "TOTAL" else DATA_FONT
        _usd_cell(ws, r, 2, proc)
        _usd_cell(ws, r, 3, cost)
        _usd_cell(ws, r, 4, gl)
        r += 1

    _auto(ws)


async def _build_form_8949_v4(wb, session, year, term, tab_name, run_id):
    ws = wb.create_sheet(title=tab_name)
    box = "B" if term == "short" else "D"
    label = "Short-Term" if term == "short" else "Long-Term"

    ws.cell(row=1, column=1, value=f"Form 8949 — {label} Capital Gains and Losses (v4)").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Tax Year: {year} | Box {box} (basis NOT reported to IRS)").font = NOTE_FONT

    r = 4
    headers = [
        "(a) Description", "(b) Date Acquired", "(c) Date Sold",
        "(d) Proceeds", "(e) Cost Basis",
        "(f) Code", "(g) Adjustment",
        "(h) Gain or Loss", "Asset", "Wallet", "Exchange", "Holding Days"
    ]
    _hdr(ws, r, headers)
    r += 1

    run_filter = "AND run_id = :rid" if run_id else ""
    params = {"y": year, "t": term}
    if run_id:
        params["rid"] = run_id

    res = await session.execute(text(f"""
        SELECT description, date_acquired, date_sold, proceeds, cost_basis,
               adjustment_code, adjustment_amount, gain_loss, asset, wallet, exchange, holding_days
        FROM tax.form_8949_v4
        WHERE tax_year = :y AND term = :t {run_filter}
        ORDER BY date_sold, asset
    """), params)

    total_proceeds = D("0")
    total_cost = D("0")
    total_gl = D("0")
    count = 0

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0])
        _data_cell(ws, r, 2, row_data[1])
        _data_cell(ws, r, 3, row_data[2])
        _usd_cell(ws, r, 4, row_data[3])
        _usd_cell(ws, r, 5, row_data[4])
        _data_cell(ws, r, 6, row_data[5] or "")
        _usd_cell(ws, r, 7, row_data[6])
        _usd_cell(ws, r, 8, row_data[7])
        _data_cell(ws, r, 9, row_data[8])
        _data_cell(ws, r, 10, row_data[9])
        _data_cell(ws, r, 11, row_data[10])
        _data_cell(ws, r, 12, row_data[11])
        total_proceeds += D(str(row_data[3] or 0))
        total_cost += D(str(row_data[4] or 0))
        total_gl += D(str(row_data[7] or 0))
        count += 1
        r += 1

    r += 1
    ws.cell(row=r, column=1, value=f"TOTALS ({count} disposals)").font = BOLD_FONT
    _usd_cell(ws, r, 4, str(total_proceeds))
    _usd_cell(ws, r, 5, str(total_cost))
    _usd_cell(ws, r, 8, str(total_gl))
    _auto(ws)


async def _build_income_schedule_v4(wb, session, year, run_id):
    ws = wb.create_sheet(title="Income Schedule")
    ws.cell(row=1, column=1, value=f"Ordinary Income Schedule (v4) — {year}").font = TITLE_FONT

    r = 4
    _hdr(ws, r, ["Date", "Type", "Asset", "Wallet", "Amount", "FMV (USD)", "Status"])
    r += 1

    run_filter = "AND run_id = :rid" if run_id else ""
    params = {"y": year}
    if run_id:
        params["rid"] = run_id

    res = await session.execute(text(f"""
        SELECT dominion_at, income_type, asset, wallet, quantity, total_fmv_usd, review_status
        FROM tax.income_events_v4
        WHERE EXTRACT(YEAR FROM dominion_at) = :y {run_filter}
        ORDER BY dominion_at ASC
    """), params)

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0], is_date=True)
        _data_cell(ws, r, 2, (row_data[1] or "").replace("_", " ").title())
        _data_cell(ws, r, 3, row_data[2])
        _data_cell(ws, r, 4, row_data[3])
        _data_cell(ws, r, 5, row_data[4])
        _usd_cell(ws, r, 6, row_data[5])
        _data_cell(ws, r, 7, row_data[6])
        r += 1

    _auto(ws)


async def _build_transfer_recon_v4(wb, session, run_id):
    ws = wb.create_sheet(title="Transfer Recon")
    ws.cell(row=1, column=1, value="Transfer Carryover Records (v4)").font = TITLE_FONT

    r = 4
    _hdr(ws, r, ["Date", "Asset", "Qty", "From Wallet", "To Wallet",
                  "Original Acquired", "Carryover Basis (USD)", "Confidence"])
    r += 1

    run_filter = "WHERE run_id = :rid" if run_id else ""
    params = {"rid": run_id} if run_id else {}

    res = await session.execute(text(f"""
        SELECT transferred_at, asset, quantity, source_wallet, dest_wallet,
               original_acquired_at, carryover_basis_usd, match_confidence
        FROM tax.transfer_carryover
        {run_filter}
        ORDER BY transferred_at ASC
    """), params)

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0], is_date=True)
        _data_cell(ws, r, 2, row_data[1])
        _data_cell(ws, r, 3, row_data[2])
        _data_cell(ws, r, 4, (row_data[3] or "").upper())
        _data_cell(ws, r, 5, (row_data[4] or "").upper())
        _data_cell(ws, r, 6, row_data[5], is_date=True)
        _usd_cell(ws, r, 7, row_data[6])
        _data_cell(ws, r, 8, row_data[7])
        r += 1

    _auto(ws)


async def _build_lot_inventory_v4(wb, session, run_id):
    ws = wb.create_sheet(title="Lot Inventory")
    ws.cell(row=1, column=1, value="FIFO Lot Inventory — v4 (Wallet-Aware)").font = TITLE_FONT

    r = 4
    _hdr(ws, r, ["Asset", "Wallet", "Acquired", "Source",
                  "Original Qty", "Remaining", "Cost/Unit (USD)", "Total Cost (USD)"])
    r += 1

    run_filter = "AND run_id = :rid" if run_id else ""
    params = {"rid": run_id} if run_id else {}

    res = await session.execute(text(f"""
        SELECT asset, wallet, original_acquired_at, source_type,
               original_quantity, remaining, cost_per_unit_usd, total_cost_usd
        FROM tax.lots_v4
        WHERE remaining > 0 {run_filter}
        ORDER BY asset, original_acquired_at
    """), params)

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0])
        _data_cell(ws, r, 2, (row_data[1] or "").upper())
        _data_cell(ws, r, 3, row_data[2], is_date=True)
        _data_cell(ws, r, 4, (row_data[3] or "").replace("_", " ").title())
        _data_cell(ws, r, 5, row_data[4])
        _data_cell(ws, r, 6, row_data[5])
        _usd_cell(ws, r, 7, row_data[6])
        _usd_cell(ws, r, 8, row_data[7])
        r += 1

    _auto(ws)


async def _build_exceptions_tab(wb, session, run_id):
    ws = wb.create_sheet(title="Exceptions")
    ws.cell(row=1, column=1, value="Open Exceptions").font = TITLE_FONT

    r = 4
    _hdr(ws, r, ["Severity", "Category", "Message", "Resolution Status", "Tax Year", "Created"])
    r += 1

    run_filter = "AND run_id = :rid" if run_id else ""
    params = {"rid": run_id} if run_id else {}

    res = await session.execute(text(f"""
        SELECT severity, category, message, resolution_status, affected_tax_year, created_at
        FROM tax.exceptions
        WHERE resolution_status = 'open' {run_filter}
        ORDER BY severity DESC, created_at ASC
    """), params)

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0])
        _data_cell(ws, r, 2, row_data[1])
        _data_cell(ws, r, 3, row_data[2])
        _data_cell(ws, r, 4, row_data[3])
        _data_cell(ws, r, 5, row_data[4])
        _data_cell(ws, r, 6, row_data[5], is_date=True)
        r += 1

    _auto(ws)


async def _build_exchange_pnl_tab(wb, session, year, run_id):
    ws = wb.create_sheet(title="Exchange P&L Summary")
    ws.cell(row=1, column=1, value=f"Realized P&L by Exchange — {year}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = NOTE_FONT

    r = 4
    _hdr(ws, r, ["Exchange", "Disposals", "Total Proceeds", "Total Basis",
                  "ST Proceeds", "ST Basis", "ST Net",
                  "LT Proceeds", "LT Basis", "LT Net", "Total Net"])
    r += 1

    run_filter = "AND run_id = :rid" if run_id else ""
    params = {"y": year}
    if run_id:
        params["rid"] = run_id

    res = await session.execute(text(f"""
        SELECT exchange, COUNT(*),
            COALESCE(SUM(proceeds), 0), COALESCE(SUM(cost_basis), 0),
            COALESCE(SUM(CASE WHEN term='short' THEN proceeds END), 0),
            COALESCE(SUM(CASE WHEN term='short' THEN cost_basis END), 0),
            COALESCE(SUM(CASE WHEN term='short' THEN gain_loss END), 0),
            COALESCE(SUM(CASE WHEN term='long' THEN proceeds END), 0),
            COALESCE(SUM(CASE WHEN term='long' THEN cost_basis END), 0),
            COALESCE(SUM(CASE WHEN term='long' THEN gain_loss END), 0),
            COALESCE(SUM(gain_loss), 0)
        FROM tax.form_8949_v4
        WHERE tax_year = :y {run_filter}
        GROUP BY exchange ORDER BY exchange
    """), params)

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, (row_data[0] or "").upper())
        _data_cell(ws, r, 2, row_data[1])
        for ci in range(3, 12):
            _usd_cell(ws, r, ci, row_data[ci - 1])
        r += 1

    _auto(ws)


async def _build_funding_flows_tab(wb, session, year):
    ws = wb.create_sheet(title="Funding Flows")
    ws.cell(row=1, column=1, value=f"Classified Funding Flows — {year or 'All'}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = NOTE_FONT

    r = 4
    _hdr(ws, r, ["Exchange", "External In (USD)", "External Out (USD)",
                  "Net External", "Internal In", "Internal Out",
                  "Income (USD)", "Unclassified"])
    r += 1

    where = ""
    params = {}
    if year:
        where = "WHERE EXTRACT(YEAR FROM event_at) = :y"
        params["y"] = year

    res = await session.execute(text(f"""
        SELECT exchange,
            COALESCE(SUM(CASE WHEN flow_class='EXTERNAL_DEPOSIT' THEN total_usd END), 0),
            COALESCE(SUM(CASE WHEN flow_class='EXTERNAL_WITHDRAWAL' THEN total_usd END), 0),
            COALESCE(SUM(CASE WHEN flow_class='EXTERNAL_DEPOSIT' THEN total_usd ELSE 0 END)
                - SUM(CASE WHEN flow_class='EXTERNAL_WITHDRAWAL' THEN total_usd ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN flow_class='INTERNAL_TRANSFER_IN' THEN total_usd END), 0),
            COALESCE(SUM(CASE WHEN flow_class='INTERNAL_TRANSFER_OUT' THEN total_usd END), 0),
            COALESCE(SUM(CASE WHEN flow_class='INCOME_RECEIPT' THEN total_usd END), 0),
            COALESCE(SUM(CASE WHEN flow_class='UNCLASSIFIED' THEN total_usd END), 0)
        FROM tax.classified_flows
        {where}
        GROUP BY exchange ORDER BY exchange
    """), params)

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, (row_data[0] or "").upper())
        for ci in range(2, 9):
            _usd_cell(ws, r, ci, row_data[ci - 1])
        r += 1

    _auto(ws)


async def _build_data_coverage_tab(wb, session):
    ws = wb.create_sheet(title="Data Coverage")
    ws.cell(row=1, column=1, value="Data Coverage & API Retention Gaps").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = NOTE_FONT

    r = 4
    _hdr(ws, r, ["Exchange", "Data Type", "API Earliest", "API Latest",
                  "CSV Earliest", "CSV Latest", "Has Gap", "Requires CSV", "CSV Imported"])
    r += 1

    res = await session.execute(text("""
        SELECT exchange, data_type, api_earliest, api_latest,
               csv_earliest, csv_latest, has_gap, requires_csv, csv_imported
        FROM tax.data_coverage ORDER BY exchange, data_type
    """))

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, (row_data[0] or "").upper())
        _data_cell(ws, r, 2, row_data[1])
        _data_cell(ws, r, 3, row_data[2], is_date=True)
        _data_cell(ws, r, 4, row_data[3], is_date=True)
        _data_cell(ws, r, 5, row_data[4], is_date=True)
        _data_cell(ws, r, 6, row_data[5], is_date=True)
        _data_cell(ws, r, 7, "YES" if row_data[6] else "NO")
        _data_cell(ws, r, 8, "YES" if row_data[7] else "NO")
        _data_cell(ws, r, 9, "YES" if row_data[8] else "NO")
        r += 1

    # CSV import history
    r += 2
    ws.cell(row=r, column=1, value="CSV Import History").font = H2_FONT
    r += 1
    _hdr(ws, r, ["Exchange", "Type", "Filename", "Rows", "Imported", "Duplicates", "Imported At"])
    r += 1

    res = await session.execute(text("""
        SELECT exchange, data_type, filename, row_count, imported_count,
               duplicate_count, imported_at
        FROM tax.csv_imports ORDER BY imported_at DESC
    """))

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, (row_data[0] or "").upper())
        _data_cell(ws, r, 2, row_data[1])
        _data_cell(ws, r, 3, row_data[2])
        _data_cell(ws, r, 4, row_data[3])
        _data_cell(ws, r, 5, row_data[4])
        _data_cell(ws, r, 6, row_data[5])
        _data_cell(ws, r, 7, row_data[6], is_date=True)
        r += 1

    _auto(ws)


async def _build_valuation_audit_tab(wb, session, year, run_id):
    ws = wb.create_sheet(title="Valuation Audit")
    ws.cell(row=1, column=1, value=f"Valuation Audit Trail — {year}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = NOTE_FONT

    r = 4
    _hdr(ws, r, ["Asset", "Event Date", "Price Date", "Price (USD)",
                  "Source", "Granularity", "Estimated?", "Manual?", "Fallback Reason"])
    r += 1

    run_filter = "AND run_id = :rid" if run_id else ""
    params = {"y": year}
    if run_id:
        params["rid"] = run_id

    res = await session.execute(text(f"""
        SELECT asset, event_at, price_date, price_usd,
               source_name, granularity, is_estimated, is_manual, fallback_reason
        FROM tax.valuation_log
        WHERE EXTRACT(YEAR FROM event_at) = :y {run_filter}
        ORDER BY event_at ASC
        LIMIT 5000
    """), params)

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0])
        _data_cell(ws, r, 2, row_data[1], is_date=True)
        _data_cell(ws, r, 3, row_data[2])
        _usd_cell(ws, r, 4, row_data[3])
        _data_cell(ws, r, 5, row_data[4])
        _data_cell(ws, r, 6, row_data[5])
        _data_cell(ws, r, 7, "YES" if row_data[6] else "NO")
        _data_cell(ws, r, 8, "YES" if row_data[7] else "NO")
        _data_cell(ws, r, 9, row_data[8] or "")
        r += 1

    _auto(ws)


async def _build_filing_readiness(wb, session, year, run_id):
    """Add FILING READINESS section to the Summary sheet."""
    ws = wb["Summary"]
    # Find the next available row
    max_row = ws.max_row + 3

    ws.cell(row=max_row, column=1, value="FILING READINESS").font = TITLE_FONT
    max_row += 1

    # Check blocking exceptions
    res = await session.execute(text("""
        SELECT COUNT(*) FROM tax.exceptions
        WHERE blocks_filing = TRUE AND resolution_status = 'open'
            AND (affected_tax_year = :y OR affected_tax_year IS NULL)
    """), {"y": year})
    blocking = res.scalar() or 0

    # Check warnings
    res = await session.execute(text("""
        SELECT COUNT(*) FROM tax.exceptions
        WHERE severity = 'WARNING' AND resolution_status = 'open'
            AND (affected_tax_year = :y OR affected_tax_year IS NULL)
    """), {"y": year})
    warnings = res.scalar() or 0

    # Check data coverage gaps
    res = await session.execute(text("""
        SELECT COUNT(*) FROM tax.data_coverage
        WHERE has_gap = TRUE AND csv_imported = FALSE
    """))
    coverage_gaps = res.scalar() or 0

    # Check missing prices
    res = await session.execute(text("""
        SELECT COUNT(*) FROM tax.exceptions
        WHERE category = 'MISSING_PRICE' AND resolution_status = 'open'
            AND (affected_tax_year = :y OR affected_tax_year IS NULL)
    """), {"y": year})
    missing_prices = res.scalar() or 0

    filing_ready = (blocking == 0 and coverage_gaps == 0 and missing_prices == 0)

    fr_label = "YES" if filing_ready else "NO"
    fr_color = "006100" if filing_ready else "CC0000"

    ws.cell(row=max_row, column=1, value="Filing Ready:").font = BOLD_FONT
    ws.cell(row=max_row, column=2, value=fr_label).font = Font(name="Arial", bold=True, size=14, color=fr_color)
    max_row += 1

    ws.cell(row=max_row, column=1, value="Blocking Issues:").font = DATA_FONT
    ws.cell(row=max_row, column=2, value=blocking)
    max_row += 1

    ws.cell(row=max_row, column=1, value="Warnings:").font = DATA_FONT
    ws.cell(row=max_row, column=2, value=warnings)
    max_row += 1

    ws.cell(row=max_row, column=1, value="Coverage Gaps:").font = DATA_FONT
    ws.cell(row=max_row, column=2, value=coverage_gaps)
    max_row += 1

    ws.cell(row=max_row, column=1, value="Missing Prices:").font = DATA_FONT
    ws.cell(row=max_row, column=2, value=missing_prices)
    max_row += 1

    if not filing_ready:
        reasons = []
        if blocking > 0:
            reasons.append(f"{blocking} blocking exception(s)")
        if coverage_gaps > 0:
            reasons.append(f"{coverage_gaps} data coverage gap(s) without CSV")
        if missing_prices > 0:
            reasons.append(f"{missing_prices} missing price(s)")
        ws.cell(row=max_row, column=1, value="Reason(s):").font = DATA_FONT
        ws.cell(row=max_row, column=2, value="; ".join(reasons))


async def _build_run_manifest_tab(wb, session, run_id):
    ws = wb.create_sheet(title="Run Manifest")
    ws.cell(row=1, column=1, value="Computation Run History").font = TITLE_FONT

    r = 4
    _hdr(ws, r, ["Run ID", "Type", "Tax Year", "Method", "Status",
                  "Started", "Completed", "Events", "Disposals", "Filing Ready"])
    r += 1

    run_filter = "WHERE id = :rid" if run_id else ""
    params = {"rid": run_id} if run_id else {}

    res = await session.execute(text(f"""
        SELECT id, run_type, tax_year, basis_method, status,
               started_at, completed_at, total_events, total_disposals, filing_ready
        FROM tax.run_manifest
        {run_filter}
        ORDER BY started_at DESC
        LIMIT 50
    """), params)

    for row_data in res.fetchall():
        _data_cell(ws, r, 1, row_data[0])
        _data_cell(ws, r, 2, row_data[1])
        _data_cell(ws, r, 3, row_data[2])
        _data_cell(ws, r, 4, row_data[3])
        _data_cell(ws, r, 5, row_data[4])
        _data_cell(ws, r, 6, row_data[5], is_date=True)
        _data_cell(ws, r, 7, row_data[6], is_date=True)
        _data_cell(ws, r, 8, row_data[7])
        _data_cell(ws, r, 9, row_data[8])
        _data_cell(ws, r, 10, row_data[9])
        r += 1

    _auto(ws)
