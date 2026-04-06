"""
Format-Aware Import Subsystem — fills exchange API retention gaps.

Supports:
  - MEXC deposit/withdrawal XLSX exports (official format)
  - NonKYC deposit/withdrawal CSV exports (official format)
  - MEXC trade CSV (API-format)
  - Generic/manual CSV with user-provided column mapping

Detects file format by extension + header fingerprinting.
Deduplicates against existing records by (exchange, exchange_id).
Records import metadata: source filename, SHA256 checksum, row count, import timestamp.
"""
import csv
import hashlib
import logging
import os
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation

from sqlalchemy import text

logger = logging.getLogger("tax-collector.csv-importer")

D = Decimal

# ── Header Fingerprints ──────────────────────────────────────────────────

MEXC_DEPOSIT_XLSX_HEADERS = ["UID", "Status", "Time", "Crypto", "Network",
                              "Deposit Amount", "TxID", "Progress"]

MEXC_WITHDRAWAL_XLSX_HEADERS = ["UID", "Status", "Time", "Crypto", "Network",
                                 "Request Amount", "Withdrawal Address", "memo",
                                 "TxID", "Trading Fee", "Settlement Amount",
                                 "Withdrawal Descriptions"]

NONKYC_DEPOSIT_CSV_HEADERS = ["Type", "Time", "Ticker", "Amount", "ValueUsd",
                               "Confirmations", "TransactionId", "Address",
                               "isPosted", "isReversed", "isSecurityConfirm"]

NONKYC_WITHDRAWAL_CSV_HEADERS = ["Type", "Time", "Ticker", "Amount", "ValueUsd",
                                  "Address", "TransactionId", "Status"]

MEXC_TRADE_CSV_HEADERS = ["symbol", "orderId", "id", "price", "qty", "quoteQty",
                           "commission", "commissionAsset", "time", "isBuyer"]


def detect_format(filepath: str) -> tuple[str, str, str]:
    """Auto-detect file format by extension + header fingerprinting.

    Returns (exchange, data_type, parser_name) or raises ValueError.
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()

        if headers == MEXC_DEPOSIT_XLSX_HEADERS:
            return ("mexc", "deposits", "mexc_deposit_xlsx")
        if headers == MEXC_WITHDRAWAL_XLSX_HEADERS:
            return ("mexc", "withdrawals", "mexc_withdrawal_xlsx")
        raise ValueError(f"Unknown XLSX format: headers={headers}")

    elif ext in (".csv", ".tsv"):
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            headers = [h.strip() for h in headers]

        if headers == NONKYC_DEPOSIT_CSV_HEADERS:
            return ("nonkyc", "deposits", "nonkyc_deposit_csv")
        if headers == NONKYC_WITHDRAWAL_CSV_HEADERS:
            return ("nonkyc", "withdrawals", "nonkyc_withdrawal_csv")
        if headers == MEXC_TRADE_CSV_HEADERS:
            return ("mexc", "trades", "mexc_trade_csv")
        raise ValueError(f"Unknown CSV format: headers={headers}")

    else:
        raise ValueError(f"Unsupported file extension: {ext}")


class CSVImporter:
    """Import exchange history from CSV/XLSX files to close API retention gaps."""

    # ── Helpers ───────────────────────────────────────────────────────────

    def _file_hash(self, filepath: str) -> str:
        h = hashlib.sha256()
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                h.update(chunk)
        return h.hexdigest()

    def _parse_ts(self, val) -> datetime | None:
        if val is None or val == "":
            return None
        if isinstance(val, datetime):
            if val.tzinfo is None:
                return val.replace(tzinfo=timezone.utc)
            return val
        if isinstance(val, (int, float)):
            return datetime.fromtimestamp(val / 1000, tz=timezone.utc)
        s = str(val).strip()
        # NonKYC format: "3/10/2026, 9:57:14 PM"
        try:
            return datetime.strptime(s, "%m/%d/%Y, %I:%M:%S %p").replace(tzinfo=timezone.utc)
        except (ValueError, TypeError):
            pass
        # MEXC XLSX format: "2026-03-11 02:03:21"
        try:
            return datetime.strptime(s, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
        except (ValueError, TypeError):
            pass
        # ISO format
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00"))
        except (ValueError, TypeError):
            pass
        # Epoch ms
        try:
            return datetime.fromtimestamp(float(s) / 1000, tz=timezone.utc)
        except (ValueError, TypeError):
            return None

    def _safe_decimal(self, val) -> str:
        try:
            return str(D(str(val or "0")))
        except (InvalidOperation, ValueError):
            return "0"

    async def _check_duplicate(self, session, table: str, exchange: str, exchange_id: str) -> bool:
        r = await session.execute(text(f"""
            SELECT 1 FROM tax.{table} WHERE exchange = :ex AND exchange_id = :eid LIMIT 1
        """), {"ex": exchange, "eid": exchange_id})
        return r.fetchone() is not None

    async def _record_import(self, session, exchange: str, data_type: str,
                             filename: str, file_hash: str, row_count: int,
                             imported: int, duplicates: int, errors: int,
                             date_start: datetime | None, date_end: datetime | None):
        await session.execute(text("""
            INSERT INTO tax.csv_imports
                (exchange, data_type, filename, file_hash, row_count, imported_count,
                 duplicate_count, error_count, date_range_start, date_range_end)
            VALUES (:ex, :dt, :fn, :fh, :rc, :ic, :dc, :ec, :ds, :de)
        """), {
            "ex": exchange, "dt": data_type, "fn": os.path.basename(filename),
            "fh": file_hash, "rc": row_count, "ic": imported, "dc": duplicates,
            "ec": errors, "ds": date_start, "de": date_end,
        })

    def _track_dates(self, ts, date_start, date_end):
        if ts:
            if date_start is None or ts < date_start:
                date_start = ts
            if date_end is None or ts > date_end:
                date_end = ts
        return date_start, date_end

    # ── Auto-detect Import ────────────────────────────────────────────────

    async def import_file(self, session, filepath: str,
                          exchange: str = None, data_type: str = None) -> dict:
        """Auto-detect format and route to the correct parser.

        Optional exchange/data_type overrides skip auto-detection.
        """
        if not exchange or not data_type:
            detected_exchange, detected_type, parser_name = detect_format(filepath)
            exchange = exchange or detected_exchange
            data_type = data_type or detected_type
        else:
            # Manual override — construct parser name
            ext = os.path.splitext(filepath)[1].lower()
            parser_name = f"{exchange}_{data_type}_{ext.lstrip('.')}"

        dispatch = {
            "mexc_deposit_xlsx": self.import_mexc_deposits_xlsx,
            "mexc_withdrawal_xlsx": self.import_mexc_withdrawals_xlsx,
            "nonkyc_deposit_csv": self.import_nonkyc_deposits_csv,
            "nonkyc_withdrawal_csv": self.import_nonkyc_withdrawals_csv,
            "mexc_trade_csv": self.import_mexc_trades,
        }

        handler = dispatch.get(parser_name)
        if handler:
            return await handler(session, filepath)

        # Fallback to generic
        return await self.import_generic(session, filepath, exchange, data_type, {})

    # ── MEXC Deposit XLSX ─────────────────────────────────────────────────

    async def import_mexc_deposits_xlsx(self, session, filepath: str) -> dict:
        """Import MEXC deposit XLSX with official headers."""
        import openpyxl
        file_hash = self._file_hash(filepath)
        imported, duplicates, errors, row_count = 0, 0, 0, 0
        date_start, date_end = None, None

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
        wb.close()

        for row in rows:
            if not row or all(c is None for c in row):
                continue
            row_count += 1
            try:
                uid, status, time_val, crypto, network, amount, txid, progress = row[:8]
                txid_str = str(txid or "")
                eid = txid_str if txid_str else f"mexc_dep_{uid}_{row_count}"

                if await self._check_duplicate(session, "deposits", "mexc", eid):
                    duplicates += 1
                    continue

                ts = self._parse_ts(time_val)
                date_start, date_end = self._track_dates(ts, date_start, date_end)

                # Strip :N output index suffix for tx_hash matching
                tx_hash = txid_str.split(":")[0] if ":" in txid_str else txid_str

                result = await session.execute(text("""
                    INSERT INTO tax.deposits
                        (exchange, exchange_id, asset, amount, network, tx_hash,
                         status, confirmed_at, raw_data, source_type, source_file)
                    VALUES ('mexc', :eid, :asset, :amount, :net, :txh,
                            :status, :ts, :raw, 'xlsx', :sf)
                    ON CONFLICT (exchange, exchange_id) DO UPDATE SET
                        status = EXCLUDED.status,
                        confirmed_at = COALESCE(EXCLUDED.confirmed_at, tax.deposits.confirmed_at),
                        raw_data = EXCLUDED.raw_data,
                        source_type = COALESCE(EXCLUDED.source_type, tax.deposits.source_type),
                        source_file = COALESCE(EXCLUDED.source_file, tax.deposits.source_file)
                    RETURNING id
                """), {
                    "eid": eid, "asset": str(crypto or ""),
                    "amount": self._safe_decimal(amount),
                    "net": str(network or ""), "txh": tx_hash,
                    "status": str(status or ""), "ts": ts,
                    "raw": str(dict(zip(MEXC_DEPOSIT_XLSX_HEADERS, row))),
                    "sf": os.path.basename(filepath),
                })
                if result.fetchone():
                    imported += 1
            except Exception as e:
                logger.warning(f"MEXC deposit XLSX row {row_count} error: {e}")
                errors += 1

        await self._record_import(session, "mexc", "deposits", filepath, file_hash,
                                  row_count, imported, duplicates, errors,
                                  date_start, date_end)
        return {"imported": imported, "duplicates": duplicates, "errors": errors,
                "row_count": row_count, "file_hash": file_hash}

    # ── MEXC Withdrawal XLSX ──────────────────────────────────────────────

    async def import_mexc_withdrawals_xlsx(self, session, filepath: str) -> dict:
        """Import MEXC withdrawal XLSX with official headers."""
        import openpyxl
        file_hash = self._file_hash(filepath)
        imported, duplicates, errors, row_count = 0, 0, 0, 0
        date_start, date_end = None, None

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        for row in rows:
            if not row or all(c is None for c in row):
                continue
            row_count += 1
            try:
                (uid, status, time_val, crypto, network, req_amount,
                 address, memo, txid, trading_fee, settlement_amount,
                 descriptions) = row[:12]
                txid_str = str(txid or "")
                eid = txid_str if txid_str else f"mexc_wd_{uid}_{row_count}"

                if await self._check_duplicate(session, "withdrawals", "mexc", eid):
                    duplicates += 1
                    continue

                ts = self._parse_ts(time_val)
                date_start, date_end = self._track_dates(ts, date_start, date_end)

                # Validate settlement = request - fee
                try:
                    req = D(str(req_amount or "0"))
                    fee = D(str(trading_fee or "0"))
                    settle = D(str(settlement_amount or "0"))
                    if settle > 0 and abs(req - fee - settle) > D("0.01"):
                        logger.warning(
                            f"MEXC withdrawal settlement mismatch: "
                            f"{req} - {fee} = {req - fee} != {settle} (row {row_count})")
                except (InvalidOperation, TypeError):
                    pass

                result = await session.execute(text("""
                    INSERT INTO tax.withdrawals
                        (exchange, exchange_id, asset, amount, fee, fee_asset,
                         network, tx_hash, address, status, confirmed_at,
                         raw_data, source_type, source_file)
                    VALUES ('mexc', :eid, :asset, :amount, :fee, :fee_asset,
                            :net, :txh, :addr, :status, :ts,
                            :raw, 'xlsx', :sf)
                    ON CONFLICT (exchange, exchange_id) DO UPDATE SET
                        status = EXCLUDED.status,
                        fee = COALESCE(EXCLUDED.fee, tax.withdrawals.fee),
                        fee_asset = COALESCE(EXCLUDED.fee_asset, tax.withdrawals.fee_asset),
                        confirmed_at = COALESCE(EXCLUDED.confirmed_at, tax.withdrawals.confirmed_at),
                        raw_data = EXCLUDED.raw_data,
                        source_type = COALESCE(EXCLUDED.source_type, tax.withdrawals.source_type),
                        source_file = COALESCE(EXCLUDED.source_file, tax.withdrawals.source_file)
                    RETURNING id
                """), {
                    "eid": eid, "asset": str(crypto or ""),
                    "amount": self._safe_decimal(req_amount),
                    "fee": self._safe_decimal(trading_fee),
                    "fee_asset": str(crypto or ""),  # MEXC fees in same asset
                    "net": str(network or ""), "txh": txid_str,
                    "addr": str(address or ""), "status": str(status or ""),
                    "ts": ts,
                    "raw": str(dict(zip(MEXC_WITHDRAWAL_XLSX_HEADERS, row))),
                    "sf": os.path.basename(filepath),
                })
                if result.fetchone():
                    imported += 1
            except Exception as e:
                logger.warning(f"MEXC withdrawal XLSX row {row_count} error: {e}")
                errors += 1

        await self._record_import(session, "mexc", "withdrawals", filepath, file_hash,
                                  row_count, imported, duplicates, errors,
                                  date_start, date_end)
        return {"imported": imported, "duplicates": duplicates, "errors": errors,
                "row_count": row_count, "file_hash": file_hash}

    # ── NonKYC Deposit CSV ────────────────────────────────────────────────

    async def import_nonkyc_deposits_csv(self, session, filepath: str) -> dict:
        """Import NonKYC deposit CSV with official headers."""
        file_hash = self._file_hash(filepath)
        imported, duplicates, errors, row_count = 0, 0, 0, 0
        date_start, date_end = None, None

        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                row_count += 1
                try:
                    eid = row.get("TransactionId", "")
                    if not eid:
                        eid = f"nonkyc_dep_{row_count}"

                    if await self._check_duplicate(session, "deposits", "nonkyc", eid):
                        duplicates += 1
                        continue

                    ts = self._parse_ts(row.get("Time"))
                    date_start, date_end = self._track_dates(ts, date_start, date_end)

                    is_posted = str(row.get("isPosted", "")).lower() == "true"

                    result = await session.execute(text("""
                        INSERT INTO tax.deposits
                            (exchange, exchange_id, asset, amount, amount_usd,
                             tx_hash, address, status, confirmed_at,
                             raw_data, source_type, source_file)
                        VALUES ('nonkyc', :eid, :asset, :amount, :amount_usd,
                                :txh, :addr, :status, :ts,
                                :raw, 'csv', :sf)
                        ON CONFLICT (exchange, exchange_id) DO UPDATE SET
                            amount_usd = COALESCE(EXCLUDED.amount_usd, tax.deposits.amount_usd),
                            status = EXCLUDED.status,
                            confirmed_at = COALESCE(EXCLUDED.confirmed_at, tax.deposits.confirmed_at),
                            raw_data = EXCLUDED.raw_data,
                            source_type = COALESCE(EXCLUDED.source_type, tax.deposits.source_type),
                            source_file = COALESCE(EXCLUDED.source_file, tax.deposits.source_file)
                        RETURNING id
                    """), {
                        "eid": eid, "asset": row.get("Ticker", ""),
                        "amount": self._safe_decimal(row.get("Amount")),
                        "amount_usd": self._safe_decimal(row.get("ValueUsd")),
                        "txh": eid,
                        "addr": row.get("Address", ""),
                        "status": "posted" if is_posted else "pending",
                        "ts": ts, "raw": str(row),
                        "sf": os.path.basename(filepath),
                    })
                    if result.fetchone():
                        imported += 1
                except Exception as e:
                    logger.warning(f"NonKYC deposit CSV row {row_count} error: {e}")
                    errors += 1

        await self._record_import(session, "nonkyc", "deposits", filepath, file_hash,
                                  row_count, imported, duplicates, errors,
                                  date_start, date_end)
        return {"imported": imported, "duplicates": duplicates, "errors": errors,
                "row_count": row_count, "file_hash": file_hash}

    # ── NonKYC Withdrawal CSV ─────────────────────────────────────────────

    async def import_nonkyc_withdrawals_csv(self, session, filepath: str) -> dict:
        """Import NonKYC withdrawal CSV with official headers."""
        file_hash = self._file_hash(filepath)
        imported, duplicates, errors, row_count = 0, 0, 0, 0
        date_start, date_end = None, None

        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                row_count += 1
                try:
                    eid = row.get("TransactionId", "")
                    if not eid:
                        eid = f"nonkyc_wd_{row_count}"

                    if await self._check_duplicate(session, "withdrawals", "nonkyc", eid):
                        duplicates += 1
                        continue

                    ts = self._parse_ts(row.get("Time"))
                    date_start, date_end = self._track_dates(ts, date_start, date_end)

                    result = await session.execute(text("""
                        INSERT INTO tax.withdrawals
                            (exchange, exchange_id, asset, amount, amount_usd,
                             tx_hash, address, status, confirmed_at,
                             raw_data, source_type, source_file)
                        VALUES ('nonkyc', :eid, :asset, :amount, :amount_usd,
                                :txh, :addr, :status, :ts,
                                :raw, 'csv', :sf)
                        ON CONFLICT (exchange, exchange_id) DO UPDATE SET
                            amount_usd = COALESCE(EXCLUDED.amount_usd, tax.withdrawals.amount_usd),
                            status = EXCLUDED.status,
                            confirmed_at = COALESCE(EXCLUDED.confirmed_at, tax.withdrawals.confirmed_at),
                            raw_data = EXCLUDED.raw_data,
                            source_type = COALESCE(EXCLUDED.source_type, tax.withdrawals.source_type),
                            source_file = COALESCE(EXCLUDED.source_file, tax.withdrawals.source_file)
                        RETURNING id
                    """), {
                        "eid": eid, "asset": row.get("Ticker", ""),
                        "amount": self._safe_decimal(row.get("Amount")),
                        "amount_usd": self._safe_decimal(row.get("ValueUsd")),
                        "txh": eid,
                        "addr": row.get("Address", ""),
                        "status": row.get("Status", ""),
                        "ts": ts, "raw": str(row),
                        "sf": os.path.basename(filepath),
                    })
                    if result.fetchone():
                        imported += 1
                except Exception as e:
                    logger.warning(f"NonKYC withdrawal CSV row {row_count} error: {e}")
                    errors += 1

        await self._record_import(session, "nonkyc", "withdrawals", filepath, file_hash,
                                  row_count, imported, duplicates, errors,
                                  date_start, date_end)
        return {"imported": imported, "duplicates": duplicates, "errors": errors,
                "row_count": row_count, "file_hash": file_hash}

    # ── MEXC Trade CSV (API format) ───────────────────────────────────────

    async def import_mexc_trades(self, session, filepath: str) -> dict:
        """Import MEXC trade CSV. Returns {imported: N, duplicates: N, errors: N}."""
        file_hash = self._file_hash(filepath)
        imported, duplicates, errors, row_count = 0, 0, 0, 0
        date_start, date_end = None, None

        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                row_count += 1
                try:
                    eid = str(row.get("id", row.get("orderId", "")))
                    if await self._check_duplicate(session, "trades", "mexc", eid):
                        duplicates += 1
                        continue

                    ts = self._parse_ts(row.get("time"))
                    date_start, date_end = self._track_dates(ts, date_start, date_end)

                    price = self._safe_decimal(row.get("price"))
                    qty = self._safe_decimal(row.get("qty"))
                    total = self._safe_decimal(row.get("quoteQty"))
                    if total == "0":
                        total = str(D(price) * D(qty))

                    is_buyer = str(row.get("isBuyer", "")).lower() in ("true", "1", "yes")

                    result = await session.execute(text("""
                        INSERT INTO tax.trades
                            (exchange, exchange_id, market, side, price, quantity,
                             total, fee, fee_asset, executed_at, raw_data, source_type, source_file)
                        VALUES ('mexc', :eid, :market, :side, :price, :qty,
                                :total, :fee, :fee_asset, :ts, :raw, 'csv', :sf)
                        ON CONFLICT (exchange, exchange_id) DO UPDATE SET
                            raw_data = EXCLUDED.raw_data,
                            source_type = COALESCE(EXCLUDED.source_type, tax.trades.source_type),
                            source_file = COALESCE(EXCLUDED.source_file, tax.trades.source_file)
                        RETURNING id
                    """), {
                        "eid": eid, "market": row.get("symbol", ""),
                        "side": "buy" if is_buyer else "sell",
                        "price": price, "qty": qty, "total": total,
                        "fee": self._safe_decimal(row.get("commission")),
                        "fee_asset": row.get("commissionAsset", ""),
                        "ts": ts, "raw": str(row),
                        "sf": os.path.basename(filepath),
                    })
                    if result.fetchone():
                        imported += 1
                except Exception as e:
                    logger.warning(f"Row {row_count} error: {e}")
                    errors += 1

        await self._record_import(session, "mexc", "trades", filepath, file_hash,
                                  row_count, imported, duplicates, errors,
                                  date_start, date_end)
        return {"imported": imported, "duplicates": duplicates, "errors": errors,
                "row_count": row_count, "file_hash": file_hash}

    # ── Legacy aliases ────────────────────────────────────────────────────

    async def import_mexc_deposits(self, session, filepath: str) -> dict:
        """Legacy alias — routes to XLSX or CSV based on extension."""
        ext = os.path.splitext(filepath)[1].lower()
        if ext in (".xlsx", ".xls"):
            return await self.import_mexc_deposits_xlsx(session, filepath)
        return await self._import_mexc_deposits_csv(session, filepath)

    async def import_mexc_withdrawals(self, session, filepath: str) -> dict:
        """Legacy alias — routes to XLSX or CSV based on extension."""
        ext = os.path.splitext(filepath)[1].lower()
        if ext in (".xlsx", ".xls"):
            return await self.import_mexc_withdrawals_xlsx(session, filepath)
        return await self._import_mexc_withdrawals_csv(session, filepath)

    async def _import_mexc_deposits_csv(self, session, filepath: str) -> dict:
        """Import MEXC deposit CSV (API-format columns)."""
        file_hash = self._file_hash(filepath)
        imported, duplicates, errors, row_count = 0, 0, 0, 0
        date_start, date_end = None, None

        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                row_count += 1
                try:
                    eid = str(row.get("txId", row.get("id", f"mexc_dep_{row_count}")))
                    if await self._check_duplicate(session, "deposits", "mexc", eid):
                        duplicates += 1
                        continue

                    ts = self._parse_ts(row.get("insertTime", row.get("completeTime")))
                    date_start, date_end = self._track_dates(ts, date_start, date_end)

                    result = await session.execute(text("""
                        INSERT INTO tax.deposits
                            (exchange, exchange_id, asset, amount, network, tx_hash,
                             status, confirmed_at, raw_data, source_type, source_file)
                        VALUES ('mexc', :eid, :asset, :amount, :net, :txh,
                                :status, :ts, :raw, 'csv', :sf)
                        ON CONFLICT (exchange, exchange_id) DO UPDATE SET
                            raw_data = EXCLUDED.raw_data,
                            source_type = COALESCE(EXCLUDED.source_type, tax.deposits.source_type),
                            source_file = COALESCE(EXCLUDED.source_file, tax.deposits.source_file)
                        RETURNING id
                    """), {
                        "eid": eid, "asset": row.get("coin", ""),
                        "amount": self._safe_decimal(row.get("amount")),
                        "net": row.get("network", ""), "txh": row.get("txId", ""),
                        "status": row.get("status", ""), "ts": ts,
                        "raw": str(row), "sf": os.path.basename(filepath),
                    })
                    if result.fetchone():
                        imported += 1
                except Exception as e:
                    logger.warning(f"Deposit row {row_count} error: {e}")
                    errors += 1

        await self._record_import(session, "mexc", "deposits", filepath, file_hash,
                                  row_count, imported, duplicates, errors,
                                  date_start, date_end)
        return {"imported": imported, "duplicates": duplicates, "errors": errors,
                "row_count": row_count, "file_hash": file_hash}

    async def _import_mexc_withdrawals_csv(self, session, filepath: str) -> dict:
        """Import MEXC withdrawal CSV (API-format columns)."""
        file_hash = self._file_hash(filepath)
        imported, duplicates, errors, row_count = 0, 0, 0, 0
        date_start, date_end = None, None

        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                row_count += 1
                try:
                    eid = str(row.get("txId", row.get("id", f"mexc_wd_{row_count}")))
                    if await self._check_duplicate(session, "withdrawals", "mexc", eid):
                        duplicates += 1
                        continue

                    ts = self._parse_ts(row.get("completeTime", row.get("applyTime")))
                    date_start, date_end = self._track_dates(ts, date_start, date_end)

                    result = await session.execute(text("""
                        INSERT INTO tax.withdrawals
                            (exchange, exchange_id, asset, amount, fee, fee_asset,
                             network, tx_hash, status, confirmed_at,
                             raw_data, source_type, source_file)
                        VALUES ('mexc', :eid, :asset, :amount, :fee, :fee_asset,
                                :net, :txh, :status, :ts,
                                :raw, 'csv', :sf)
                        ON CONFLICT (exchange, exchange_id) DO UPDATE SET
                            raw_data = EXCLUDED.raw_data,
                            source_type = COALESCE(EXCLUDED.source_type, tax.withdrawals.source_type),
                            source_file = COALESCE(EXCLUDED.source_file, tax.withdrawals.source_file)
                        RETURNING id
                    """), {
                        "eid": eid, "asset": row.get("coin", ""),
                        "amount": self._safe_decimal(row.get("amount")),
                        "fee": self._safe_decimal(row.get("transactionFee")),
                        "fee_asset": row.get("coin", ""),
                        "net": row.get("network", ""), "txh": row.get("txId", ""),
                        "status": row.get("status", ""), "ts": ts,
                        "raw": str(row), "sf": os.path.basename(filepath),
                    })
                    if result.fetchone():
                        imported += 1
                except Exception as e:
                    logger.warning(f"Withdrawal row {row_count} error: {e}")
                    errors += 1

        await self._record_import(session, "mexc", "withdrawals", filepath, file_hash,
                                  row_count, imported, duplicates, errors,
                                  date_start, date_end)
        return {"imported": imported, "duplicates": duplicates, "errors": errors,
                "row_count": row_count, "file_hash": file_hash}

    # ── Generic Import ────────────────────────────────────────────────────

    async def import_generic(self, session, filepath: str, exchange: str,
                             data_type: str, column_map: dict) -> dict:
        """Import any CSV with a user-provided column mapping."""
        file_hash = self._file_hash(filepath)
        imported, duplicates, errors, row_count = 0, 0, 0, 0
        date_start, date_end = None, None

        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                row_count += 1
                try:
                    mapped = {}
                    for target, source in column_map.items():
                        mapped[target] = row.get(source, "")

                    eid = mapped.get("exchange_id", f"{exchange}_{data_type}_{row_count}")
                    table = "trades" if data_type == "trades" else data_type
                    if await self._check_duplicate(session, table, exchange, eid):
                        duplicates += 1
                        continue

                    ts_val = mapped.get("timestamp") or mapped.get("time")
                    ts = self._parse_ts(ts_val)
                    date_start, date_end = self._track_dates(ts, date_start, date_end)

                    if data_type == "trades":
                        result = await session.execute(text("""
                            INSERT INTO tax.trades
                                (exchange, exchange_id, market, side, price, quantity,
                                 total, fee, fee_asset, executed_at, raw_data, source_type, source_file)
                            VALUES (:ex, :eid, :market, :side, :price, :qty,
                                    :total, :fee, :fee_asset, :ts, :raw, 'csv', :sf)
                            ON CONFLICT (exchange, exchange_id) DO NOTHING
                            RETURNING id
                        """), {
                            "ex": exchange, "eid": eid,
                            "market": mapped.get("market", ""),
                            "side": mapped.get("side", ""),
                            "price": self._safe_decimal(mapped.get("price")),
                            "qty": self._safe_decimal(mapped.get("quantity")),
                            "total": self._safe_decimal(mapped.get("total")),
                            "fee": self._safe_decimal(mapped.get("fee")),
                            "fee_asset": mapped.get("fee_asset", ""),
                            "ts": ts, "raw": str(row),
                            "sf": os.path.basename(filepath),
                        })
                        if result.fetchone():
                            imported += 1
                    elif data_type == "deposits":
                        result = await session.execute(text("""
                            INSERT INTO tax.deposits
                                (exchange, exchange_id, asset, amount, tx_hash, address,
                                 status, confirmed_at, raw_data, source_type, source_file)
                            VALUES (:ex, :eid, :asset, :amount, :txh, :addr,
                                    :status, :ts, :raw, 'csv', :sf)
                            ON CONFLICT (exchange, exchange_id) DO NOTHING
                            RETURNING id
                        """), {
                            "ex": exchange, "eid": eid,
                            "asset": mapped.get("asset", ""),
                            "amount": self._safe_decimal(mapped.get("amount")),
                            "txh": mapped.get("tx_hash", ""),
                            "addr": mapped.get("address", ""),
                            "status": mapped.get("status", ""),
                            "ts": ts, "raw": str(row),
                            "sf": os.path.basename(filepath),
                        })
                        if result.fetchone():
                            imported += 1
                    elif data_type == "withdrawals":
                        result = await session.execute(text("""
                            INSERT INTO tax.withdrawals
                                (exchange, exchange_id, asset, amount, fee, tx_hash, address,
                                 status, confirmed_at, raw_data, source_type, source_file)
                            VALUES (:ex, :eid, :asset, :amount, :fee, :txh, :addr,
                                    :status, :ts, :raw, 'csv', :sf)
                            ON CONFLICT (exchange, exchange_id) DO NOTHING
                            RETURNING id
                        """), {
                            "ex": exchange, "eid": eid,
                            "asset": mapped.get("asset", ""),
                            "amount": self._safe_decimal(mapped.get("amount")),
                            "fee": self._safe_decimal(mapped.get("fee", "0")),
                            "txh": mapped.get("tx_hash", ""),
                            "addr": mapped.get("address", ""),
                            "status": mapped.get("status", ""),
                            "ts": ts, "raw": str(row),
                            "sf": os.path.basename(filepath),
                        })
                        if result.fetchone():
                            imported += 1
                except Exception as e:
                    logger.warning(f"Generic row {row_count} error: {e}")
                    errors += 1

        await self._record_import(session, exchange, data_type, filepath, file_hash,
                                  row_count, imported, duplicates, errors,
                                  date_start, date_end)
        return {"imported": imported, "duplicates": duplicates, "errors": errors,
                "row_count": row_count, "file_hash": file_hash}
